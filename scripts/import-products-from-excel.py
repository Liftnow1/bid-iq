"""Import / round-trip the master products xlsx back to the DB.

Reads the edited products-master-*.xlsx and diffs against the live products
table. Applies INSERTs (new rows w/ no id), UPDATEs (id present, fields
changed), and DELETEs (ids in DB that are missing from the edited file)
inside one transaction. Prints a preview and prompts for confirmation
unless --yes is passed.

Usage:
    python scripts/import-products-from-excel.py --file path/to/edited.xlsx
    python scripts/import-products-from-excel.py --file ... --yes
    python scripts/import-products-from-excel.py --file ... --dry-run
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parent.parent


def _load_dotenv() -> None:
    env_path = REPO_ROOT / ".env"
    if not env_path.exists(): return
    pat = re.compile(r"\s*([A-Z_][A-Z0-9_]*)\s*=\s*(.+?)\s*$")
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            m = pat.match(line)
            if m and not os.environ.get(m.group(1)):
                os.environ[m.group(1)] = m.group(2)


_load_dotenv()
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


VALID_CATEGORIES = {
    "two-post-lift", "four-post-lift", "scissor-lift", "mobile-column",
    "light-duty-inground", "heavy-duty-inground", "vertical-rise-lift",
    "parallelogram-lift", "low-rise-lift", "rolling-jack", "unclassified",
}
VALID_STATUSES = {"current", "discontinued", "unknown"}


def parse_int(v) -> int | None:
    if v is None or v == "": return None
    try: return int(v)
    except (TypeError, ValueError): return None


def parse_variants(v) -> list[str]:
    if v is None: return []
    if isinstance(v, list): return [str(x).strip() for x in v if str(x).strip()]
    s = str(v).strip()
    if not s: return []
    return [x.strip() for x in s.split(",") if x.strip()]


def read_xlsx(path: Path) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Products" not in wb.sheetnames:
        raise RuntimeError(f"sheet 'Products' not found in {path.name}; sheets={wb.sheetnames}")
    ws = wb["Products"]
    headers = [c.value for c in ws[1]]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        # Skip fully-empty rows (Excel sometimes leaves trailing blanks).
        if not any(c is not None and str(c).strip() for c in r):
            continue
        row = dict(zip(headers, r))
        # Coerce types and clean whitespace
        for k, v in list(row.items()):
            if isinstance(v, str): row[k] = v.strip() or None
        out.append({
            "id": parse_int(row.get("id")),
            "brand": row.get("brand"),
            "sku": row.get("sku"),
            "family_name": row.get("family_name"),
            "product_name": row.get("product_name"),
            "description": row.get("description"),
            "category": row.get("category") or "unclassified",
            "capacity_lbs": parse_int(row.get("capacity_lbs")),
            "variant_skus": parse_variants(row.get("variant_skus")),
            "status": row.get("status") or "unknown",
            "is_ali_certified": (str(row.get("is_ali_cert") or "N").upper() == "Y"),
            "ali_cert_date": row.get("ali_cert_date"),
            "source": row.get("source") or "unknown",
            "source_file": row.get("source_file"),
            "notes": row.get("notes"),
        })
    return out


def fetch_db_rows():
    import psycopg
    with psycopg.connect(os.environ["DATABASE_URL"]) as conn, conn.cursor() as cur:
        cur.execute("""
          SELECT p.id, b.name AS brand, p.sku, p.family_name, p.product_name,
                 p.description, p.category, p.capacity_lbs,
                 coalesce(p.variant_skus, '[]'::jsonb) AS variant_skus,
                 p.status, p.is_ali_certified, p.ali_cert_date,
                 p.source, p.source_file, p.notes
          FROM products p JOIN brands b ON b.id = p.brand_id
        """)
        cols = [d[0] for d in cur.description]
        rows = []
        for r in cur.fetchall():
            d = dict(zip(cols, r))
            v = d.get("variant_skus") or []
            if isinstance(v, str):
                try: v = json.loads(v)
                except Exception: v = []
            d["variant_skus"] = v if isinstance(v, list) else []
            rows.append(d)
        return {r["id"]: r for r in rows}


def fields_differ(xrow: dict, dbrow: dict) -> list[str]:
    """List of column names that differ between xlsx row and DB row."""
    diffs = []
    for k in [
        "brand", "sku", "family_name", "product_name", "description",
        "category", "capacity_lbs", "status", "is_ali_certified",
        "source", "source_file", "notes",
    ]:
        a = xrow.get(k)
        b = dbrow.get(k)
        # Normalize None vs "" so trivial blank diffs don't flag.
        if a == "" : a = None
        if b == "" : b = None
        if a != b:
            diffs.append(k)
    # variant_skus comparison: order-insensitive
    a = sorted(xrow.get("variant_skus") or [])
    b = sorted(dbrow.get("variant_skus") or [])
    if a != b:
        diffs.append("variant_skus")
    # ali_cert_date comparison — DB returns date object; xlsx comes as str
    a = xrow.get("ali_cert_date")
    b = dbrow.get("ali_cert_date")
    if a is not None and hasattr(b, "isoformat"):
        if str(a) != b.isoformat(): diffs.append("ali_cert_date")
    elif a is not None or b is not None:
        if str(a) != (b.isoformat() if b and hasattr(b, "isoformat") else b):
            diffs.append("ali_cert_date")
    return diffs


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True, help="path to edited xlsx")
    ap.add_argument("--yes", action="store_true", help="skip confirmation")
    ap.add_argument("--dry-run", action="store_true",
                    help="show diff but do not apply")
    ap.add_argument("--no-deletes", action="store_true",
                    help="apply inserts/updates only; don't delete rows missing from xlsx")
    args = ap.parse_args()

    xpath = Path(args.file)
    if not xpath.exists():
        print(f"file not found: {xpath}", file=sys.stderr)
        return 2

    xrows = read_xlsx(xpath)
    print(f"[import] xlsx rows: {len(xrows)}")
    db_by_id = fetch_db_rows()
    print(f"[import] db rows: {len(db_by_id)}")

    # Diff
    inserts = []   # rows w/o id
    updates = []   # rows with matching id but fields differ
    nochanges = 0
    seen_ids = set()
    missing_id_in_db = []  # id in xlsx but not in DB (probably wrong id; treat as new)

    for r in xrows:
        rid = r.get("id")
        if rid is None:
            inserts.append(r)
            continue
        if rid not in db_by_id:
            missing_id_in_db.append(r)
            continue
        seen_ids.add(rid)
        diffs = fields_differ(r, db_by_id[rid])
        if diffs:
            updates.append((r, diffs))
        else:
            nochanges += 1

    deletes = [db_by_id[rid] for rid in db_by_id if rid not in seen_ids]

    print()
    print(f"  no-change:  {nochanges}")
    print(f"  updates:    {len(updates)}")
    print(f"  inserts:    {len(inserts)}")
    print(f"  unknown id: {len(missing_id_in_db)} (treated as inserts)")
    print(f"  deletes:    {len(deletes)}")
    print()

    # Validate
    bad = []
    for r in xrows:
        if not r.get("brand") or not r.get("sku"):
            bad.append(("missing brand or sku", r))
        if r.get("category") not in VALID_CATEGORIES:
            bad.append((f"invalid category {r.get('category')!r}", r))
        if r.get("status") not in VALID_STATUSES:
            bad.append((f"invalid status {r.get('status')!r}", r))
    if bad:
        print(f"[import] FOUND {len(bad)} VALIDATION ERRORS — aborting:")
        for msg, r in bad[:10]:
            print(f"  {msg}  brand={r.get('brand')!r} sku={r.get('sku')!r}")
        return 2

    # Show samples
    if updates:
        print("Sample updates (up to 5):")
        for r, diffs in updates[:5]:
            print(f"  id={r['id']} brand={r['brand']} sku={r['sku']} → changed: {','.join(diffs)}")
        print()
    if deletes:
        print(f"Pending DELETEs (showing all {len(deletes)}):")
        for d in deletes:
            print(f"  id={d['id']} brand={d['brand']:12s} sku={d['sku']:25s} {d.get('family_name','')[:40]}")
        print()
    if inserts or missing_id_in_db:
        all_new = inserts + missing_id_in_db
        print(f"Sample inserts (up to 5 of {len(all_new)}):")
        for r in all_new[:5]:
            print(f"  brand={r['brand']:12s} sku={r['sku']:25s} cat={r['category']:18s}")
        print()

    if args.dry_run:
        print("[import] DRY-RUN — no changes applied.")
        return 0

    if not args.yes:
        prompt = f"Apply: {len(updates)} updates, {len(inserts) + len(missing_id_in_db)} inserts, {0 if args.no_deletes else len(deletes)} deletes? [y/N] "
        ans = input(prompt).strip().lower()
        if ans != "y":
            print("[import] cancelled.")
            return 1

    import psycopg
    with psycopg.connect(os.environ["DATABASE_URL"], autocommit=False) as conn:
        with conn.cursor() as cur:
            # Resolve brand_id per brand-name (cache)
            brand_ids: dict[str, int] = {}
            def bid(name: str) -> int:
                if name in brand_ids: return brand_ids[name]
                cur.execute("SELECT id FROM brands WHERE lower(name) = lower(%s)", (name,))
                row = cur.fetchone()
                if row:
                    brand_ids[name] = row[0]
                    return row[0]
                cur.execute("INSERT INTO brands (name) VALUES (%s) RETURNING id", (name,))
                brand_ids[name] = cur.fetchone()[0]
                return brand_ids[name]

            # DELETEs
            if not args.no_deletes and deletes:
                ids_to_delete = [d["id"] for d in deletes]
                cur.execute("DELETE FROM products WHERE id = ANY(%s)", (ids_to_delete,))
                print(f"[import] deleted {len(ids_to_delete)} rows")

            # UPDATEs
            for r, _ in updates:
                ali_date = r.get("ali_cert_date") or None
                cur.execute(
                    """
                    UPDATE products SET
                      brand_id=%s, sku=%s, family_name=%s, product_name=%s,
                      description=%s, category=%s, capacity_lbs=%s,
                      variant_skus=%s::jsonb, status=%s, is_ali_certified=%s,
                      ali_cert_date=%s, source=%s, source_file=%s, notes=%s,
                      updated_at=NOW()
                    WHERE id=%s
                    """,
                    (
                        bid(r["brand"]), r["sku"], r.get("family_name"),
                        r.get("product_name"), r.get("description"),
                        r["category"], r.get("capacity_lbs"),
                        json.dumps(r.get("variant_skus") or []),
                        r["status"], r.get("is_ali_certified"),
                        ali_date, r.get("source"), r.get("source_file"),
                        r.get("notes"), r["id"],
                    ),
                )
            if updates: print(f"[import] updated {len(updates)} rows")

            # INSERTs (rows with no id, plus rows whose id wasn't found in DB)
            for r in inserts + missing_id_in_db:
                ali_date = r.get("ali_cert_date") or None
                cur.execute(
                    """
                    INSERT INTO products (
                      brand_id, sku, family_name, product_name, description,
                      category, capacity_lbs, variant_skus, status,
                      is_ali_certified, ali_cert_date, source, source_file, notes
                    ) VALUES (
                      %s, %s, %s, %s, %s,
                      %s, %s, %s::jsonb, %s,
                      %s, %s, %s, %s, %s
                    )
                    ON CONFLICT (brand_id, sku) DO UPDATE SET
                      family_name=EXCLUDED.family_name,
                      product_name=EXCLUDED.product_name,
                      description=EXCLUDED.description,
                      category=EXCLUDED.category,
                      capacity_lbs=EXCLUDED.capacity_lbs,
                      variant_skus=EXCLUDED.variant_skus,
                      status=EXCLUDED.status,
                      is_ali_certified=EXCLUDED.is_ali_certified,
                      ali_cert_date=EXCLUDED.ali_cert_date,
                      source=EXCLUDED.source,
                      source_file=EXCLUDED.source_file,
                      notes=EXCLUDED.notes,
                      updated_at=NOW()
                    """,
                    (
                        bid(r["brand"]), r["sku"], r.get("family_name"),
                        r.get("product_name"), r.get("description"),
                        r["category"], r.get("capacity_lbs"),
                        json.dumps(r.get("variant_skus") or []),
                        r["status"], r.get("is_ali_certified"),
                        ali_date, r.get("source"), r.get("source_file"),
                        r.get("notes"),
                    ),
                )
            if inserts or missing_id_in_db:
                print(f"[import] inserted/upserted {len(inserts)+len(missing_id_in_db)} rows")

        conn.commit()
        print("[import] committed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
