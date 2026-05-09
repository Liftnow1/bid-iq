"""Iter15 price-sheet parser.

Reads a manufacturer price sheet (xlsx or pdf) for a single brand, extracts
the lift product list (no pricing, no accessories), groups variants under
their family product, and either prints a CSV for review (--dry-run) or
upserts into the `products` table.

Usage:
    python scripts/parse-price-sheet.py \
        --brand challenger \
        --file "data/pillar2-staging/Challenger Lifts - Current SW 04-20-2026.xlsx" \
        [--sheet "Sheet1"] \
        [--dry-run] \
        [--out logs/price-sheet-challenger.csv]

How it works:

1. Read the file preserving structure (xlsx via openpyxl with merged-cell
   awareness; PDF via pdfplumber tables, falling back to text).
2. Hand the rows to Claude with a structured-output prompt that says:
   "extract families + variants for vehicle lifts and rolling jacks only;
    skip accessories, options, parts, kits; do not include pricing".
3. Parse Claude's JSON response into product rows.
4. Print a per-brand CSV for review and (unless --dry-run) upsert into
   the `products` table with status='current' and source='price-sheet'.

The `variant_skus` JSONB column on `products` carries the configured
variants; queries can hit them via the GIN index.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from pathlib import Path
from typing import Any, Optional

REPO_ROOT = Path(__file__).resolve().parent.parent


def _load_dotenv() -> None:
    env_path = REPO_ROOT / ".env"
    if not env_path.exists():
        return
    pat = re.compile(r"\s*([A-Z_][A-Z0-9_]*)\s*=\s*(.+?)\s*$")
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            m = pat.match(line)
            if m and not os.environ.get(m.group(1)):
                os.environ[m.group(1)] = m.group(2)


_load_dotenv()
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")


# Categories must match the products.category CHECK constraint.
VALID_CATEGORIES = [
    "two-post-lift", "four-post-lift", "scissor-lift", "mobile-column",
    "light-duty-inground", "heavy-duty-inground", "vertical-rise-lift",
    "parallelogram-lift", "low-rise-lift", "rolling-jack", "unclassified",
]


def read_xlsx_rows(path: Path, sheet_name: Optional[str] = None) -> list[list[str]]:
    """Return the sheet as a list of cell-string rows. Empty cells become ''."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} not found in {path.name}. "
                f"Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([
            "" if v is None else str(v).strip()
            for v in row
        ])
    return rows


def read_pdf_rows(path: Path) -> list[list[str]]:
    """Return PDF tables as flat row lists. Falls back to text-line splitting
    if pdfplumber can't find structured tables."""
    import pdfplumber
    rows: list[list[str]] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            if tables:
                for tbl in tables:
                    for row in tbl:
                        rows.append([
                            "" if c is None else str(c).strip()
                            for c in row
                        ])
                continue
            # No tables on this page — extract text and split each line on
            # whitespace so it at least looks row-shaped to the LLM.
            text = page.extract_text() or ""
            for ln in text.splitlines():
                cells = re.split(r"\s{2,}", ln.strip())
                if cells:
                    rows.append(cells)
    return rows


def rows_to_text_block(rows: list[list[str]], max_chars: int = 50_000) -> str:
    """Compact a row list into a pipe-delimited text block bounded by max_chars.
    Empty rows are dropped; long values are truncated to keep the prompt small."""
    out: list[str] = []
    total = 0
    for r in rows:
        if not any(c.strip() for c in r):
            continue
        cells = [(c[:120] + "...") if len(c) > 120 else c for c in r]
        line = " | ".join(cells)
        if total + len(line) + 1 > max_chars:
            out.append(f"... [truncated; original had {len(rows)} rows]")
            break
        out.append(line)
        total += len(line) + 1
    return "\n".join(out)


SYSTEM_PROMPT = """You are extracting a vehicle-lift product catalog from a manufacturer's price sheet for a structured database. The user is Liftnow, a vehicle-lift dealer building an inventory-management product.

You will receive a price-sheet excerpt as pipe-delimited rows (the sheet's actual visual structure). Your job is to identify the MODEL FAMILIES of vehicle lifts and rolling jacks, plus the variant SKUs (configurations) under each family.

## Inclusion rules (HARD)

INCLUDE: vehicle lifts (any kind) and rolling jacks.

EXCLUDE everything else:
- accessories, kits, options, add-ons (oil drains, work platforms, jack pads, lights, alignment kits, casters, dollies, etc.)
- replacement parts, service parts, spare parts
- shipping/freight charges
- installation services
- price-sheet section headers, totals, subtotals, footnotes
- compressors, tire changers, balancers, alignment racks (unless the brand is explicitly a lift maker AND the row IS a lift)
- "removed", "discontinued", "discontinued -" rows, replacement-only listings

If a row is ambiguous (e.g. "Optional kit" but not clearly a lift), EXCLUDE it.

## Family-level grouping

A FAMILY is one distinct lift model. Variants are configurations of the SAME physical lift differing by color, length, mount type, finish, voltage, runway type, etc.

Read the sheet's structure to identify families:
- Section headers ("01 Coats 2 Post Lifts", "Mobile Column Lifts") often label a family
- Rows with the same model number prefix are usually variants of one family
- Description text often gives the family name verbatim ("RX16 Series", "MP18 Mobile Column")

When in doubt, prefer FEWER families with MORE variants over many tiny families.

## Categories (must use one of these exact strings)

- two-post-lift
- four-post-lift
- scissor-lift
- mobile-column         (wheel-engaging/mobile-column/portable-column lifts)
- light-duty-inground
- heavy-duty-inground   (>30,000 lb capacity inground lifts)
- vertical-rise-lift    (vertical-rise short-stroke lifts)
- parallelogram-lift
- low-rise-lift         (low/mid-rise pad/frame lifts under 36" rise)
- rolling-jack
- unclassified          (only if you really can't tell)

## Output format

Return STRICT JSON (no prose, no markdown). Shape:

{
  "products": [
    {
      "sku": "RX16",                  // family-level base SKU; the canonical handle for the family
      "family_name": "Hunter RX16",   // human-readable family name
      "product_name": "RX16 Series Two-Post Lift",  // full product name
      "description": "16,000 lb capacity 2-post lift with...",
      "category": "two-post-lift",
      "capacity_lbs": 16000,           // null if not stated
      "variant_skus": ["RX16-1AT-LP", "RX16-1AT-FX", "RX16-1AT-CG"],
      "notes": "optional human notes"  // can be null/omitted
    },
    ...
  ]
}

Rules for the JSON:
- DO NOT include pricing in any field
- variant_skus is an array of strings; if there are no separate variants, use an empty array []
- capacity_lbs is integer pounds (e.g. 16000), not "16K" or "16,000 lbs"
- if the sheet excerpt only contains accessories/parts and no lifts, return {"products": []}
- if you can't determine the family vs variant grouping for a row, put the row's SKU in the family `sku` field with empty variant_skus
"""


def build_user_prompt(
    brand: str, file_label: str, rows_block: str, guidance: str = ""
) -> str:
    guidance_block = ""
    if guidance.strip():
        guidance_block = f"""
## Brand-specific guidance (HARD — overrides any conflicting general rules)

{guidance.strip()}
"""
    return f"""Brand: {brand}
Source file: {file_label}
{guidance_block}
Price-sheet excerpt (pipe-delimited rows):

{rows_block}

Extract families + variants per the system rules. Vehicle lifts and rolling jacks only. No pricing. JSON only."""


def call_claude(brand: str, file_label: str, rows_block: str, guidance: str = "") -> dict:
    import anthropic
    client = anthropic.Anthropic()
    # 16K is enough for ~200 family entries with full variant arrays. The
    # original 8K cap truncated mid-JSON on BendPak (798 source rows) and
    # PKS (PDF text). Sonnet 4.5 supports up to 64K output tokens; staying
    # at 16K keeps per-call cost reasonable while preventing truncation
    # for any plausible single price-sheet brand.
    resp = client.messages.create(
        model="claude-sonnet-4-5",
        max_tokens=16384,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": build_user_prompt(brand, file_label, rows_block, guidance)}],
    )
    text = "".join(b.text for b in resp.content if b.type == "text").strip()
    # Be defensive — strip any code-fence Claude sometimes wraps despite "JSON only".
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```\s*$", "", text)
    try:
        return json.loads(text)
    except json.JSONDecodeError as e:
        # Save the raw response for debugging
        debug_path = REPO_ROOT / "logs" / f"price-sheet-{brand}-raw.txt"
        debug_path.parent.mkdir(parents=True, exist_ok=True)
        debug_path.write_text(text, encoding="utf-8")
        raise RuntimeError(
            f"Claude returned non-JSON; raw saved to {debug_path}: {e}"
        )


def normalize_product(p: dict, brand: str, source_file: str) -> dict:
    """Coerce one product dict to the shape we'll insert into the products table."""
    sku = (p.get("sku") or "").strip()
    if not sku:
        return {}
    cat = p.get("category") or "unclassified"
    if cat not in VALID_CATEGORIES:
        cat = "unclassified"
    cap = p.get("capacity_lbs")
    if cap is not None:
        try:
            cap = int(cap)
            if cap < 0 or cap > 1_000_000:
                cap = None
        except (TypeError, ValueError):
            cap = None
    variants = p.get("variant_skus") or []
    if not isinstance(variants, list):
        variants = []
    variants = [str(v).strip() for v in variants if v and str(v).strip()]
    return {
        "brand": brand,
        "sku": sku,
        "family_name": (p.get("family_name") or None),
        "product_name": (p.get("product_name") or None),
        "description": (p.get("description") or None),
        "category": cat,
        "capacity_lbs": cap,
        "variant_skus": variants,
        "status": "current",
        "source": "price-sheet",
        "source_file": source_file,
        "notes": (p.get("notes") or None),
    }


def upsert_products(rows: list[dict]) -> tuple[int, int]:
    """Insert/update one row per brand+sku. Returns (inserted, updated)."""
    import psycopg
    inserted = 0
    updated = 0
    with psycopg.connect(os.environ["DATABASE_URL"], autocommit=True) as conn, conn.cursor() as cur:
        # Resolve brand_id once per brand
        brand_ids: dict[str, int] = {}
        for r in rows:
            b = r["brand"]
            if b in brand_ids:
                continue
            cur.execute("SELECT id FROM brands WHERE lower(name) = lower(%s)", (b,))
            row = cur.fetchone()
            if row:
                brand_ids[b] = row[0]
            else:
                cur.execute(
                    "INSERT INTO brands (name) VALUES (%s) ON CONFLICT DO NOTHING RETURNING id",
                    (b,),
                )
                row = cur.fetchone()
                if row:
                    brand_ids[b] = row[0]
                else:
                    cur.execute("SELECT id FROM brands WHERE lower(name) = lower(%s)", (b,))
                    brand_ids[b] = cur.fetchone()[0]

        for r in rows:
            bid = brand_ids[r["brand"]]
            cur.execute(
                """
                INSERT INTO products (
                  brand_id, sku, family_name, product_name, description,
                  category, capacity_lbs, variant_skus, status, source,
                  source_file, notes, updated_at
                )
                VALUES (
                  %s, %s, %s, %s, %s,
                  %s, %s, %s::jsonb, %s, %s,
                  %s, %s, NOW()
                )
                ON CONFLICT (brand_id, sku)
                DO UPDATE SET
                  family_name = EXCLUDED.family_name,
                  product_name = EXCLUDED.product_name,
                  description = EXCLUDED.description,
                  category = EXCLUDED.category,
                  capacity_lbs = EXCLUDED.capacity_lbs,
                  variant_skus = EXCLUDED.variant_skus,
                  status = EXCLUDED.status,
                  source = EXCLUDED.source,
                  source_file = EXCLUDED.source_file,
                  notes = EXCLUDED.notes,
                  updated_at = NOW()
                RETURNING (xmax = 0) AS inserted
                """,
                (
                    bid, r["sku"], r["family_name"], r["product_name"], r["description"],
                    r["category"], r["capacity_lbs"], json.dumps(r["variant_skus"]),
                    r["status"], r["source"], r["source_file"], r["notes"],
                ),
            )
            row = cur.fetchone()
            if row and row[0]:
                inserted += 1
            else:
                updated += 1
    return inserted, updated


def write_csv(out_path: Path, rows: list[dict]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "brand", "sku", "family_name", "product_name", "description",
        "category", "capacity_lbs", "variant_skus_count", "variant_skus",
        "status", "source", "source_file", "notes",
    ]
    with open(out_path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            row = {**r, "variant_skus_count": len(r.get("variant_skus") or [])}
            row["variant_skus"] = ", ".join(r.get("variant_skus") or [])
            w.writerow(row)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--brand", required=True)
    ap.add_argument("--file", required=True, help="path to xlsx or pdf")
    ap.add_argument("--sheet", default=None, help="xlsx sheet name (default: first)")
    ap.add_argument("--dry-run", action="store_true",
                    help="print CSV preview only; do not write to DB")
    ap.add_argument("--out", default=None,
                    help="CSV output path (default: logs/price-sheet-<brand>.csv)")
    ap.add_argument("--guidance", default="",
                    help="brand-specific guidance text to inject into the prompt")
    ap.add_argument("--guidance-file", default=None,
                    help="path to a text file containing the brand-specific guidance")
    ap.add_argument("--wipe-brand", action="store_true",
                    help="DELETE existing rows for this brand before upserting (for full restructures)")
    args = ap.parse_args()

    guidance = args.guidance
    if args.guidance_file:
        guidance = (Path(args.guidance_file).read_text(encoding="utf-8")).strip()

    fpath = Path(args.file)
    if not fpath.exists():
        print(f"file not found: {fpath}", file=sys.stderr)
        return 2

    print(f"[parser] brand={args.brand} file={fpath.name}")
    if fpath.suffix.lower() in (".xlsx", ".xlsm"):
        rows = read_xlsx_rows(fpath, args.sheet)
    elif fpath.suffix.lower() == ".pdf":
        rows = read_pdf_rows(fpath)
    else:
        print(f"unsupported file type: {fpath.suffix}", file=sys.stderr)
        return 2
    print(f"[parser] read {len(rows)} rows from {fpath.suffix}")

    rows_block = rows_to_text_block(rows)
    print(f"[parser] rows_block size: {len(rows_block):,} chars")

    if guidance:
        print(f"[parser] guidance ({len(guidance)} chars): {guidance[:120]}{'...' if len(guidance)>120 else ''}")
    print(f"[parser] calling Claude...")
    result = call_claude(args.brand, fpath.name, rows_block, guidance)
    raw_products = result.get("products", [])
    print(f"[parser] Claude returned {len(raw_products)} raw products")

    products = []
    for p in raw_products:
        norm = normalize_product(p, args.brand, fpath.name)
        if norm:
            products.append(norm)
    print(f"[parser] {len(products)} normalized products")

    if not products:
        print("[parser] no products found; nothing to write")
        return 0

    out_path = Path(args.out) if args.out else (REPO_ROOT / "logs" / f"price-sheet-{args.brand}.csv")
    write_csv(out_path, products)
    print(f"[parser] wrote CSV preview: {out_path}")

    print(f"\n[parser] preview (first 5 products):")
    for p in products[:5]:
        v_count = len(p["variant_skus"])
        v_preview = ", ".join(p["variant_skus"][:3]) + ("..." if v_count > 3 else "")
        print(
            f"  sku={p['sku']:20s} cat={p['category']:18s} cap={p['capacity_lbs']!s:>6s} "
            f"  variants={v_count:3d} [{v_preview}]"
        )
        print(f"    {p['family_name']}")

    if args.dry_run:
        print(f"\n[parser] DRY-RUN — DB unchanged. Review CSV above and re-run without --dry-run.")
        return 0

    if args.wipe_brand:
        # Full-restructure path. Delete every existing row for this brand
        # before upserting the new set. Used when the family-grouping itself
        # needs to change (e.g. Hunter consolidating to RX10K/RX12K/RX14K/RX16K).
        import psycopg
        with psycopg.connect(os.environ["DATABASE_URL"], autocommit=True) as conn, conn.cursor() as cur:
            cur.execute(
                "DELETE FROM products WHERE brand_id = (SELECT id FROM brands WHERE lower(name)=lower(%s))",
                (args.brand,),
            )
            print(f"[parser] WIPED {cur.rowcount} existing rows for brand={args.brand}")

    print(f"\n[parser] upserting {len(products)} rows into products table...")
    ins, upd = upsert_products(products)
    print(f"[parser] DB: inserted={ins} updated={upd}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
