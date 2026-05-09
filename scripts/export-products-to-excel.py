"""Export the current products table to a single review-ready xlsx.

Paul reviews the export by hand, edits whatever's wrong (categories,
family groupings, capacities, names, dropping rows, adding rows), then
sends it back. The companion import script (see scripts/import-products-
from-excel.py) will diff the returned file against the live DB and apply
inserts/updates/deletes.

Output: data/pillar2-staging/products-master-vN-YYYYMMDD.xlsx

The xlsx has:
- One master sheet 'Products' with one row per family product, columns
  ordered for human readability (brand first, technical fields last).
- A frozen header row + autofilter for quick filtering / sorting.
- Data validation on the `category` column (dropdown of the 10 valid
  categories so Paul can't typo a category and break the round-trip).
- A hidden `id` column carrying the DB primary key — the import script
  uses this to detect which rows are edits vs new inserts.

Usage:
    python scripts/export-products-to-excel.py
    python scripts/export-products-to-excel.py --out path/to/file.xlsx
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent


def _load_dotenv() -> None:
    env_path = REPO_ROOT / ".env"
    if not env_path.exists():
        return
    pat = re.compile(r"\s*([A-Z_][A-Z0-9_]*)\s*=\s*(.+?)\s*$")
    with open(env_path, encoding="utf-8") as f:
        for line in f:
            m = pat.match(line)
            if m and not os.environ.get(m.group(1)):
                os.environ[m.group(1)] = m.group(2)


_load_dotenv()
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")


VALID_CATEGORIES = [
    "two-post-lift", "four-post-lift", "scissor-lift", "mobile-column",
    "light-duty-inground", "heavy-duty-inground", "vertical-rise-lift",
    "parallelogram-lift", "low-rise-lift", "rolling-jack", "unclassified",
]
VALID_STATUSES = ["current", "discontinued", "unknown"]

# Column order: human-readable first, technical / round-trip fields at
# the right. The hidden `id` column lives at column A so the import
# script can find it without relying on header order.
COLUMNS = [
    ("id",           10,  True),   # hidden — DB primary key for round-trip
    ("brand",        14,  False),
    ("sku",          22,  False),
    ("family_name",  36,  False),
    ("product_name", 44,  False),
    ("description",  60,  False),
    ("category",     20,  False),
    ("capacity_lbs", 12,  False),
    ("variant_count", 8,  False),
    ("variant_skus", 60,  False),
    ("status",       12,  False),
    ("is_ali_cert",  10,  False),
    ("ali_cert_date",12,  False),
    ("source",       14,  False),
    ("source_file",  40,  False),
    ("notes",        40,  False),
]


def fetch_products():
    import psycopg
    with psycopg.connect(os.environ["DATABASE_URL"]) as conn, conn.cursor() as cur:
        cur.execute(
            """
            SELECT p.id, b.name AS brand, p.sku, p.family_name, p.product_name,
                   p.description, p.category, p.capacity_lbs,
                   coalesce(p.variant_skus, '[]'::jsonb) AS variant_skus,
                   p.status, p.is_ali_certified, p.ali_cert_date,
                   p.source, p.source_file, p.notes
            FROM products p JOIN brands b ON b.id = p.brand_id
            ORDER BY b.name, p.category, p.capacity_lbs NULLS LAST, p.sku
            """
        )
        cols = [d[0] for d in cur.description]
        rows = []
        for r in cur.fetchall():
            row = dict(zip(cols, r))
            # variant_skus comes back as a Python list (psycopg auto-decodes
            # jsonb); coerce to string if it's somehow not a list.
            v = row.get("variant_skus") or []
            if isinstance(v, str):
                try: v = json.loads(v)
                except Exception: v = []
            row["variant_skus"] = v if isinstance(v, list) else []
            rows.append(row)
    return rows


def build_workbook(rows: list[dict], out_path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # Header row.
    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="305496")
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data rows.
    for r in rows:
        ws.append([
            r.get("id"),
            r.get("brand"),
            r.get("sku"),
            r.get("family_name") or "",
            r.get("product_name") or "",
            r.get("description") or "",
            r.get("category") or "unclassified",
            r.get("capacity_lbs") if r.get("capacity_lbs") is not None else "",
            len(r.get("variant_skus") or []),
            ", ".join(r.get("variant_skus") or []),
            r.get("status") or "unknown",
            "Y" if r.get("is_ali_certified") else "N",
            r.get("ali_cert_date").isoformat() if r.get("ali_cert_date") else "",
            r.get("source") or "unknown",
            r.get("source_file") or "",
            r.get("notes") or "",
        ])

    # Column widths + hidden flags.
    for idx, (name, width, hidden) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width
        if hidden:
            ws.column_dimensions[col_letter].hidden = True

    # Freeze header + autofilter on data range.
    ws.freeze_panes = "B2"
    last_row = len(rows) + 1
    last_col_letter = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    # Data validation: category dropdown.
    cat_col_idx = headers.index("category") + 1
    cat_letter = get_column_letter(cat_col_idx)
    dv_cat = DataValidation(
        type="list",
        formula1='"' + ",".join(VALID_CATEGORIES) + '"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid category",
        error="Use one of: " + ", ".join(VALID_CATEGORIES),
    )
    dv_cat.add(f"{cat_letter}2:{cat_letter}{last_row}")
    ws.add_data_validation(dv_cat)

    # Data validation: status dropdown.
    st_col_idx = headers.index("status") + 1
    st_letter = get_column_letter(st_col_idx)
    dv_st = DataValidation(
        type="list",
        formula1='"' + ",".join(VALID_STATUSES) + '"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Invalid status",
        error="Use one of: " + ", ".join(VALID_STATUSES),
    )
    dv_st.add(f"{st_letter}2:{st_letter}{last_row}")
    ws.add_data_validation(dv_st)

    # Validation for is_ali_cert (Y/N).
    aali_idx = headers.index("is_ali_cert") + 1
    aali_letter = get_column_letter(aali_idx)
    dv_yn = DataValidation(
        type="list",
        formula1='"Y,N"',
        allow_blank=False,
    )
    dv_yn.add(f"{aali_letter}2:{aali_letter}{last_row}")
    ws.add_data_validation(dv_yn)

    # Add an instructions sheet for Paul.
    ws2 = wb.create_sheet("README", 0)
    ws2.column_dimensions["A"].width = 100
    instructions = [
        ("Master product catalog — review sheet", True),
        ("", False),
        (f"Generated: {datetime.now(timezone.utc).isoformat()}", False),
        (f"Total rows: {len(rows)} family products across {len({r['brand'] for r in rows})} brands", False),
        ("", False),
        ("How to review this file", True),
        ("", False),
        ("1. Switch to the 'Products' tab.", False),
        ("2. The first column is a hidden 'id' (DB primary key). DO NOT edit or unhide it — the import script uses it to figure out which rows you changed vs which are new.", False),
        ("3. Edit any field in any row. Common edits:", False),
        ("    - Wrong category → use the dropdown (Excel will reject typos).", False),
        ("    - Wrong family grouping → fix `sku` / `family_name` / merge variants by deleting one row and adding its variant_skus to another's `variant_skus` list (comma-separated).", False),
        ("    - Bad description / family_name → just rewrite.", False),
        ("    - Wrong capacity_lbs → enter the integer pounds (e.g. 16000, not '16K').", False),
        ("    - Add a missing product → insert a NEW row at the bottom; leave `id` blank. Set `source` = 'manual' and any source_file you want.", False),
        ("    - Remove an accessory that slipped through → delete the row entirely.", False),
        ("4. variant_skus is a comma-separated list. The import script will split on commas, trim whitespace, and store as JSON.", False),
        ("5. is_ali_cert: Y or N (dropdown).", False),
        ("6. ali_cert_date: free-form ISO date 'YYYY-MM-DD' or leave blank.", False),
        ("7. status: 'current' (in price sheet), 'discontinued', or 'unknown'.", False),
        ("8. SAVE AS xlsx (do not change format). Send me the edited file.", False),
        ("", False),
        ("Valid categories (pick from the dropdown):", True),
        ("    " + ", ".join(VALID_CATEGORIES), False),
        ("", False),
        ("Round-trip semantics", True),
        ("    - Row with id present and unchanged fields → no-op.", False),
        ("    - Row with id present and changed fields → UPDATE that row.", False),
        ("    - Row with NO id → INSERT a new product.", False),
        ("    - Row missing from your sheet (id was in DB but not in returned file) → DELETE that row.", False),
        ("        ↳ The import script will print a list of pending DELETEs and ask for confirmation before applying.", False),
    ]
    for i, (text, is_header) in enumerate(instructions, start=1):
        cell = ws2.cell(row=i, column=1, value=text)
        if is_header:
            cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser()
    today = datetime.now().strftime("%Y%m%d")
    default_out = REPO_ROOT / "data" / "pillar2-staging" / f"products-master-v1-{today}.xlsx"
    ap.add_argument("--out", default=str(default_out),
                    help=f"output xlsx path (default: {default_out})")
    args = ap.parse_args()

    rows = fetch_products()
    print(f"[export] fetched {len(rows)} product families from DB")
    out_path = Path(args.out)
    build_workbook(rows, out_path)
    print(f"[export] wrote {out_path}")
    print(f"[export] Open in Excel, edit, save as .xlsx, send back.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
