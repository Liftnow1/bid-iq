#!/usr/bin/env python3
"""
Ingest data/pillar2-staging/SVI_Lift_Products_Catalog.xlsx into the
products + product_external_refs tables.

Rules:
  1. OUR ORIGINAL RUN takes precedence over SVI. If an SVI row fuzzy-matches
     an existing product row (same brand + SKU exact/prefix match), we DO
     NOT overwrite the product. We only insert a product_external_refs row
     pointing at the existing product so Paul keeps the SVI URLs for later
     use.
  2. SVI rows that don't match anything in our DB get inserted as new
     products with source='svi-catalog' and is_ali_certified=NULL ("Unknown"
     per Paul's instruction — equipment not on our original list defaults
     to ALI status unknown).
  3. Discontinued status is per-MAKE in the SVI Summary sheet (e.g., all 90
     AMMCO/Hennessy models are discontinued). New SVI products inherit the
     make-level status: 'discontinued' if the Summary lists "Discontinued",
     otherwise 'current'. Existing products are NEVER touched, so their
     status stays whatever our price-sheet run set.
  4. Every SVI row — whether matched or new — gets a row in
     product_external_refs so Paul can query SVI URLs later for other ideas.

Brand handling:
  - Known overlap brands (BendPak, Mohawk, Rotary, Challenger, Hunter, PKS,
    Coats, ARI/Hetra, Stertil Koni, Nussbaum) map to our existing brand_id.
  - All other SVI makes become new brand rows.

Category mapping: SVI's 20 categories collapse to our 10 + 'unclassified'
for non-lift entries (cylinders, parts catalogs).

SKU match: lowercased, hyphens/underscores/spaces stripped. Try exact
against product.sku and every variant_skus[]; fall back to a min-4-char
bidirectional prefix match.
"""

from __future__ import annotations
import argparse
import json
import os
import re
import sys
from pathlib import Path

import openpyxl  # type: ignore
import psycopg  # type: ignore
from psycopg.rows import dict_row  # type: ignore

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = REPO_ROOT / "data" / "pillar2-staging" / "SVI_Lift_Products_Catalog.xlsx"

# --- Brand aliases: SVI make name -> our brand name (lowercased) ---
# Only includes the overlaps. Everything else gets a fresh brand row.
SVI_TO_OUR_BRAND: dict[str, str] = {
    "ari/hetra": "ari-hetra",
    "bendpak": "bendpak",
    "challenger": "challenger",
    "coats": "coats",
    "hunter": "hunter",
    "mohawk": "mohawk",
    "nussbaum": "nussbaum",  # we have an empty Nussbaum brand row
    "pks lifts": "pks",
    "rotary lift": "rotary",
    "snap-on": "snap-on",    # empty brand row
    "stertil koni": "stertil-koni",
    "atlas": "atlas",
    "dannmar": "dannmar",
    "direct lift": "direct-lift",
}

# --- Category mapping: SVI -> our 10 + unclassified ---
SVI_CATEGORY_MAP: dict[str, str] = {
    "above ground 2-post lifts":                "two-post-lift",
    "above ground 4-post lifts":                "four-post-lift",
    "above ground scissor lifts":               "scissor-lift",
    "above ground mobile column lifts":         "mobile-column",
    "above ground rolling bridge jacks":        "rolling-jack",
    "above ground rolling jack lifts":          "rolling-jack",
    "above ground mid rise lifts":              "low-rise-lift",
    "above ground low rise lifts":              "low-rise-lift",
    "above ground parking lifts":               "four-post-lift",
    "above ground motorcycle lifts":            "unclassified",
    "above ground parallelogram lifts":         "parallelogram-lift",
    "above ground 1-post lifts":                "unclassified",
    "above ground alignment lift":              "four-post-lift",
    "in ground front and rear / fore and aft lifts": "heavy-duty-inground",
    "in ground single post lifts":              "light-duty-inground",
    "in ground side by side lifts":             "heavy-duty-inground",
    "in ground single post swing rail lifts":   "light-duty-inground",
    "automotive lifts (new installs)":          "unclassified",
    "automotive lift products and accessories": "unclassified",
    "hydraulic cylinders / custom lifting":     "unclassified",
}


def load_db_url() -> str:
    url = os.environ.get("DATABASE_URL") or os.environ.get("POSTGRES_URL")
    if url:
        return url
    for envfile in (".env.local", ".env"):
        ep = REPO_ROOT / envfile
        if not ep.exists():
            continue
        for line in ep.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if "=" not in line or line.startswith("#"):
                continue
            k, v = line.split("=", 1)
            if k.strip() in ("DATABASE_URL", "POSTGRES_URL"):
                return v.strip().strip('"').strip("'")
    print("ERROR: DATABASE_URL not set", file=sys.stderr)
    sys.exit(1)


def slugify_brand(name: str) -> str:
    """Turn an SVI make name into our brand slug.
    'AMMCO / Hennessy' -> 'ammco-hennessy'
    'John Bean / JBC'  -> 'john-bean-jbc'
    """
    s = name.lower().strip()
    s = re.sub(r"[\s/]+", "-", s)
    s = re.sub(r"[^a-z0-9-]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s


def normalize_sku(s: str) -> str:
    return re.sub(r"[\s\-_]+", "", (s or "").lower())


def read_xlsx() -> tuple[dict[str, bool], list[dict]]:
    """Returns (per_make_discontinued, rows).
    per_make_discontinued: {make_name: True/False}
    rows: each row is {'make', 'category', 'model', 'page_url', 'resource_urls': [...]}.
    """
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    # --- Summary sheet: per-make Discontinued flag ---
    per_make: dict[str, bool] = {}
    sw = wb["Summary"]
    in_table = False
    hdr_idx: dict[str, int] = {}
    for row in sw.iter_rows(values_only=True):
        if row and row[0] == "Manufacturer":
            hdr_idx = {str(v).strip(): i for i, v in enumerate(row) if v}
            in_table = True
            continue
        if not in_table:
            continue
        if not row or not row[0]:
            continue
        make = str(row[hdr_idx["Manufacturer"]]).strip()
        disc = str(row[hdr_idx.get("Discontinued", -1)] or "").strip().lower()
        per_make[make] = (disc == "discontinued")

    # --- SVI Lift Products sheet ---
    sp = wb["SVI Lift Products"]
    rows = []
    header_row = None
    for row in sp.iter_rows(values_only=True):
        if header_row is None:
            header_row = [str(c).strip() if c else "" for c in row]
            continue
        if not row or not row[0]:
            continue
        rec = {h: row[i] if i < len(row) else None for i, h in enumerate(header_row)}
        make = (rec.get("Make") or "").strip()
        cat = (rec.get("Category") or "").strip()
        model = (rec.get("Model") or "").strip()
        page_url = (rec.get("Model Page URL") or "").strip()
        res = []
        for k in header_row:
            if k.startswith("Resource Endpoint"):
                v = rec.get(k)
                if v:
                    res.append(str(v).strip())
        if not (make and model):
            continue
        rows.append({
            "make": make,
            "category": cat,
            "model": model,
            "page_url": page_url,
            "resource_urls": res,
        })
    wb.close()
    return per_make, rows


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args()

    per_make, rows = read_xlsx()
    print(f"SVI Summary makes:      {len(per_make)}  ({sum(per_make.values())} discontinued)")
    print(f"SVI rows to process:    {len(rows)}")

    conn = psycopg.connect(load_db_url(), autocommit=False)
    cur = conn.cursor(row_factory=dict_row)

    # --- Build current brand index ---
    cur.execute("SELECT id, lower(name) AS name_lower, name FROM brands")
    brand_by_name = {r["name_lower"]: r["id"] for r in cur.fetchall()}

    # --- Resolve each SVI make to a brand_id (create brand rows for new) ---
    make_to_brand_id: dict[str, int] = {}
    new_brands = 0
    fake_id = 10_000_000  # only used in dry-run so find_match treats them as new
    for make in per_make:
        # Map via alias table first
        slug = SVI_TO_OUR_BRAND.get(make.lower()) or slugify_brand(make)
        if slug in brand_by_name:
            make_to_brand_id[make] = brand_by_name[slug]
            continue
        # Brand doesn't exist locally yet
        if not args.dry_run:
            cur.execute(
                "INSERT INTO brands (name, relationship_type) VALUES (%s, 'unknown') "
                "ON CONFLICT (name) DO UPDATE SET name = EXCLUDED.name RETURNING id",
                (slug,),
            )
            row = cur.fetchone()
            make_to_brand_id[make] = row["id"]
        else:
            # In dry-run, give it a unique positive fake ID so the rest of
            # the logic (which gates on brand_id > 0) still runs and gives
            # accurate "new" counts.
            make_to_brand_id[make] = fake_id
            fake_id += 1
        new_brands += 1

    print(f"  brands matched to existing: {len(per_make) - new_brands}")
    print(f"  brands newly created:        {new_brands}")

    # --- Build a SKU index of EXISTING products by brand_id ---
    cur.execute("""
        SELECT p.id, p.brand_id, p.sku,
               coalesce(p.variant_skus, '[]'::jsonb) AS variant_skus
        FROM products p
    """)
    products_rows = cur.fetchall()
    # brand_id -> list of (sku_norm, product_id)
    by_brand: dict[int, list[tuple[str, int]]] = {}
    for r in products_rows:
        skus = [r["sku"]] + list(r["variant_skus"] or [])
        for s in skus:
            sn = normalize_sku(s)
            if sn:
                by_brand.setdefault(r["brand_id"], []).append((sn, r["id"]))

    def find_match(brand_id: int, model_norm: str) -> int | None:
        candidates = by_brand.get(brand_id) or []
        # exact first
        for sku_norm, pid in candidates:
            if sku_norm == model_norm:
                return pid
        # bidirectional prefix, min 4 chars
        if len(model_norm) < 4:
            return None
        for sku_norm, pid in candidates:
            if len(sku_norm) < 4:
                continue
            if sku_norm.startswith(model_norm) or model_norm.startswith(sku_norm):
                return pid
        return None

    # --- Process each SVI row ---
    stats = {
        "matched_existing": 0,
        "new_product_inserts": 0,
        "ref_inserts": 0,
        "ref_skipped_existing": 0,
        "unmapped_category": 0,
    }
    per_brand_stats: dict[str, dict[str, int]] = {}

    for row in rows:
        make = row["make"]
        brand_id = make_to_brand_id.get(make)
        if not brand_id:
            continue

        model = row["model"]
        model_norm = normalize_sku(model)
        cat = SVI_CATEGORY_MAP.get(row["category"].lower(), "unclassified")
        if cat == "unclassified" and row["category"].lower() not in SVI_CATEGORY_MAP:
            stats["unmapped_category"] += 1

        bstats = per_brand_stats.setdefault(make, {"matched": 0, "new": 0, "refs": 0})

        matched_pid = find_match(brand_id, model_norm)

        if matched_pid is None:
            # Insert NEW product. Status: discontinued if the make is
            # flagged in SVI's summary, else current.
            status = "discontinued" if per_make.get(make) else "current"
            if not args.dry_run:
                cur.execute("""
                    INSERT INTO products (brand_id, sku, family_name, category, status,
                                          source, is_ali_certified, notes)
                    VALUES (%s, %s, %s, %s, %s, 'svi-catalog', NULL, %s)
                    ON CONFLICT (brand_id, sku) DO UPDATE
                      SET notes = EXCLUDED.notes
                    RETURNING id
                """, (
                    brand_id, model, f"{make} {model}", cat, status,
                    f"Imported from SVI catalog ({row['category']})",
                ))
                matched_pid = cur.fetchone()["id"]
                # Add the new product to our SKU index so subsequent rows
                # in the same run dedupe correctly.
                by_brand.setdefault(brand_id, []).append((model_norm, matched_pid))
            stats["new_product_inserts"] += 1
            bstats["new"] += 1
        else:
            stats["matched_existing"] += 1
            bstats["matched"] += 1

        # Insert/upsert the product_external_refs row
        if not args.dry_run and matched_pid:
            try:
                cur.execute("""
                    INSERT INTO product_external_refs
                      (product_id, source, external_sku, external_make,
                       external_category, page_url, resource_urls)
                    VALUES (%s, 'svi-catalog', %s, %s, %s, %s, %s::jsonb)
                    ON CONFLICT (product_id, source, external_sku) DO UPDATE
                      SET external_make = EXCLUDED.external_make,
                          external_category = EXCLUDED.external_category,
                          page_url = EXCLUDED.page_url,
                          resource_urls = EXCLUDED.resource_urls,
                          updated_at = NOW()
                    RETURNING (xmax = 0) AS inserted
                """, (
                    matched_pid, model, make, row["category"],
                    row["page_url"], json.dumps(row["resource_urls"]),
                ))
                ins = cur.fetchone()["inserted"]
                if ins:
                    stats["ref_inserts"] += 1
                else:
                    stats["ref_skipped_existing"] += 1
            except Exception as e:
                print(f"  ref insert failed for {make} {model}: {e}", file=sys.stderr)
        else:
            stats["ref_inserts"] += 1
        bstats["refs"] += 1

    if args.dry_run:
        conn.rollback()
    else:
        conn.commit()

    print()
    print("Summary:")
    for k, v in stats.items():
        print(f"  {k:<28} {v}")
    print()
    print(f"{'make':<35} {'matched':>8} {'new':>6} {'refs':>6}")
    print("-" * 60)
    for m in sorted(per_brand_stats):
        s = per_brand_stats[m]
        print(f"{m:<35} {s['matched']:>8} {s['new']:>6} {s['refs']:>6}")

    if args.dry_run:
        print()
        print("[DRY RUN] No DB writes committed.")
    cur.close()
    conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
